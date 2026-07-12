"""
Human baseline calibration — compare LLM-as-judge scores against human reviewer scores.

Human scores come from evals/human_baseline/human_scores.xlsx (never modified here).
That workbook has three sheets:
  - human_scores: one row per agent (5 rows) — the top-level agent scores used
    by all the tests and the Summary table below.
  - PredictDelayRecords / SimulationRecords: 50 individually human-reviewed
    records each, backing the Predict / Simulate rows in human_scores as their
    mean (verified — see write_baseline_report). The other three agents
    (Diagnose, Recommend, Email) remain single-sample.

LLM scores come from the LATEST eval run (evals/reports/judge_scores_latest.json,
written by conftest.pytest_sessionfinish); the llm_* columns in the XLSX are used
only as a fallback when no eval run has been recorded.

Writes a Markdown report to evals/reports/human_baseline_report_<timestamp>.md.

Run:
    uv run pytest evals/test_eval_human_baseline.py -v
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

BASELINE_XLSX = Path(__file__).parent / "human_baseline" / "human_scores.xlsx"
REPORTS_DIR   = Path(__file__).parent / "reports"
LATEST_SCORES = REPORTS_DIR / "judge_scores_latest.json"
DIMENSIONS    = ["relevance", "faithfulness", "safety"]

# Detail sheets backing the Predict / Simulate agent-level scores with
# individually human-reviewed records (50 each). Report display is capped at
# DETAIL_DISPLAY_CAP rows per sheet — the full set lives in the workbook itself.
_DETAIL_SHEETS = {
    "Predict Delivery Delays":  "PredictDelayRecords",
    "Simulate Delay Prediction": "SimulationRecords",
}
DETAIL_DISPLAY_CAP = 5

# Keyword → judge-record agent name, tolerant of label variants in the XLS.
# ORDER MATTERS: "predict" must be checked LAST — the simulate agent's label
# ("Simulate Delay Prediction") contains the substring "predict" and would
# otherwise be mis-mapped to the predict agent's scores.
_AGENT_KEYWORDS = [
    ("simul",     "Simulate Delay Prediction"),
    ("diagnos",   "Diagnose Delay Patterns"),
    ("recommend", "Recommendation Expert Agent"),
    ("email",     "Email Alert Agent"),
    ("predict",   "Predict Delivery Delays"),
]


def _latest_llm_scores() -> tuple[dict, str]:
    """Return ({agent_name: {dim: score}}, source_label) for the LLM side.

    Source priority:
      1. In-memory judge records from THIS pytest session (when the baseline
         runs inside the full eval suite) — immune to file write-ordering.
      2. reports/judge_scores_latest.json from the most recent completed run.
      3. Empty dict — callers fall back to the XLS llm_* columns per row.
    Sources 1 and 2 are merged (session records win per agent).
    """
    session_scores: dict = {}
    try:
        import judge  # evals dir is on sys.path via conftest
        # grouped_scores collapses per-part records (e.g. predict's Summary /
        # LLM Insights sub-scores) into one averaged score per base agent name,
        # matching what conftest's report writer and judge_scores_latest.json use.
        for agent, scores in judge.grouped_scores().items():
            session_scores[agent] = {d: float(scores[d]) for d in DIMENSIONS}
    except Exception:
        pass

    file_scores: dict = {}
    file_src = ""
    if LATEST_SCORES.exists():
        payload = json.loads(LATEST_SCORES.read_text(encoding="utf-8"))
        file_scores = payload.get("scores", {})
        file_src = f"latest eval run ({payload.get('generated', 'unknown time')})"

    merged = {**file_scores, **session_scores}
    if session_scores:
        return merged, "current eval session (in-memory judge scores)"
    if file_scores:
        return merged, file_src
    return {}, "xlsx snapshot (no eval run recorded — run evals/run_evals.py first)"


def _llm_scores_for(row: dict, latest: dict) -> dict:
    """LLM scores for one XLS row: latest run if the agent matches, else llm_* columns."""
    label = (row.get("agent") or "").lower()
    for keyword, agent_name in _AGENT_KEYWORDS:
        if keyword in label and agent_name in latest:
            return {d: float(latest[agent_name][d]) for d in DIMENSIONS}
    return {d: float(row[f"llm_{d}"]) for d in DIMENSIONS}


def _load() -> list[dict]:
    """Load the agent-level summary sheet (5 rows) — the sheet all tests key off."""
    wb = openpyxl.load_workbook(str(BASELINE_XLSX), data_only=True)
    ws = wb["human_scores"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = []
    for row_cells in ws.iter_rows(min_row=2):
        vals = [c.value for c in row_cells]
        if all(v is None for v in vals):
            continue
        rows.append({
            headers[j]: str(vals[j]).strip() if vals[j] not in (None, "") else ""
            for j in range(len(headers))
        })
    return rows


def _load_detail_sheet(sheet_name: str) -> list[dict]:
    """Load one of the 50-record detail sheets (PredictDelayRecords / SimulationRecords).
    Header row is row 2 (row 1 is a merged title cell)."""
    wb = openpyxl.load_workbook(str(BASELINE_XLSX), data_only=True)
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[2]]
    rows = []
    for row_cells in ws.iter_rows(min_row=3):
        vals = [c.value for c in row_cells]
        if all(v is None for v in vals):
            continue
        row = dict(zip(headers, vals))
        if row.get("human_relevance") is None:
            continue  # unscored record — skip
        rows.append(row)
    return rows


def _scored_rows(rows: list[dict]) -> list[dict]:
    return [
        r for r in rows
        if all((r.get(f"human_{d}") or "").strip() for d in DIMENSIONS)
    ]


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def rows():
    assert BASELINE_XLSX.exists(), f"XLSX not found: {BASELINE_XLSX}"
    return _load()


@pytest.fixture(scope="module")
def scored(rows):
    s = _scored_rows(rows)
    assert s, (
        f"No rows with human scores found in {BASELINE_XLSX}. "
        "Fill in human_relevance, human_faithfulness, human_safety (1-5) for each row."
    )
    return s


# ── tests ─────────────────────────────────────────────────────────────────────

def test_all_rows_scored(rows):
    """All rows must have human scores filled in."""
    missing = [r["sample_id"] for r in rows if r not in _scored_rows(rows)]
    assert not missing, f"Missing human scores for: {', '.join(missing)}"


def test_human_scores_in_range(scored):
    """Human scores must be 1–5."""
    out_of_range = []
    for r in scored:
        for d in DIMENSIONS:
            val = float(r[f"human_{d}"])
            if not (1.0 <= val <= 5.0):
                out_of_range.append(f"{r['sample_id']}.{d}={val}")
    assert not out_of_range, f"Scores out of 1–5 range: {out_of_range}"


def test_no_large_divergence(scored):
    """Flag any agent where human mean differs from LLM mean by more than 1.5 points."""
    latest, _source = _latest_llm_scores()
    large_gaps = []
    for r in scored:
        llm        = _llm_scores_for(r, latest)
        llm_mean   = sum(llm[d] for d in DIMENSIONS) / 3
        human_mean = sum(float(r[f"human_{d}"]) for d in DIMENSIONS) / 3
        if abs(llm_mean - human_mean) > 1.5:
            large_gaps.append(f"{r['agent']}: LLM={llm_mean:.2f} Human={human_mean:.2f}")
    assert not large_gaps, f"Large LLM-human divergence (>1.5 mean pts): {large_gaps}"


def test_detail_records_match_summary(rows):
    """The Predict / Simulate rows in human_scores must equal the mean of their
    50-record detail sheets — catches drift if one is edited without the other."""
    mismatches = []
    for agent_name, sheet_name in _DETAIL_SHEETS.items():
        detail_rows = _load_detail_sheet(sheet_name)
        assert detail_rows, f"{sheet_name} has no scored records"

        avg = {d: sum(float(r[f"human_{d}"]) for r in detail_rows) / len(detail_rows) for d in DIMENSIONS}
        summary_row = next((r for r in rows if r["agent"] == agent_name), None)
        assert summary_row is not None, f"No '{agent_name}' row in human_scores sheet"

        for d in DIMENSIONS:
            summary_val = float(summary_row[f"human_{d}"])
            if abs(summary_val - avg[d]) >= 0.01:
                mismatches.append(
                    f"{agent_name}.{d}: summary={summary_val:.3f} vs "
                    f"{sheet_name} mean={avg[d]:.3f} (n={len(detail_rows)})"
                )
    assert not mismatches, f"Detail-sheet averages don't match summary row: {mismatches}"


# ── report writer ─────────────────────────────────────────────────────────────

@pytest.fixture(scope="session", autouse=True)
def write_baseline_report():
    """Write the human baseline calibration report after all tests finish."""
    yield

    if not BASELINE_XLSX.exists():
        return
    all_rows = _load()
    scored   = _scored_rows(all_rows)
    if not scored:
        return

    latest, llm_source = _latest_llm_scores()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y%m%dT%H%M%S")
    report_path = REPORTS_DIR / f"human_baseline_report_{timestamp}.md"

    lines = [
        "# Human Baseline Calibration Report",
        f"\n**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ",
        f"**Human scores:** `evals/human_baseline/human_scores.xlsx`  ",
        f"**LLM scores:** {llm_source}  ",
        f"**Samples scored:** {len(scored)} / {len(all_rows)}\n",
        "---\n",
        "## Summary\n",
        "| Agent | LLM Mean | Human Mean | Diff | Aligned |",
        "|-------|----------|------------|------|---------|",
    ]
    for r in scored:
        llm        = _llm_scores_for(r, latest)
        llm_mean   = sum(llm[d] for d in DIMENSIONS) / 3
        human_mean = sum(float(r[f"human_{d}"]) for d in DIMENSIONS) / 3
        diff       = human_mean - llm_mean
        sign       = "+" if diff > 0 else ""
        aligned    = "Yes" if abs(diff) <= 1.0 else "No — review"
        lines.append(
            f"| {r['agent']} | {llm_mean:.2f} | {human_mean:.2f} | {sign}{diff:.2f} | {aligned} |"
        )
    lines.append("")

    # ── Relevance ─────────────────────────────────────────────────────────────
    lines += [
        "---\n",
        "## 1. Relevance\n",
        "_Does the output address the task and user need?_\n",
        "| Agent | LLM | Human | Delta |",
        "|-------|-----|-------|-------|",
    ]
    for r in scored:
        lv = _llm_scores_for(r, latest)["relevance"]; hv = float(r["human_relevance"])
        dv = hv - lv; ds = "+" if dv > 0 else ""
        lines.append(f"| {r['agent']} | {lv:.1f}/5 | {hv:.0f}/5 | {ds}{dv:.1f} |")
    lines.append("")

    # ── Faithfulness ──────────────────────────────────────────────────────────
    lines += [
        "---\n",
        "## 2. Faithfulness\n",
        "_Are all claims grounded in the data — no hallucinated numbers?_\n",
        "| Agent | LLM | Human | Delta |",
        "|-------|-----|-------|-------|",
    ]
    for r in scored:
        lv = _llm_scores_for(r, latest)["faithfulness"]; hv = float(r["human_faithfulness"])
        dv = hv - lv; ds = "+" if dv > 0 else ""
        lines.append(f"| {r['agent']} | {lv:.1f}/5 | {hv:.0f}/5 | {ds}{dv:.1f} |")
    lines.append("")

    # ── Safety ────────────────────────────────────────────────────────────────
    lines += [
        "---\n",
        "## 3. Safety\n",
        "_Absence of harmful, misleading, or inappropriate content._\n",
        "| Agent | LLM | Human | Delta |",
        "|-------|-----|-------|-------|",
    ]
    for r in scored:
        lv = _llm_scores_for(r, latest)["safety"]; hv = float(r["human_safety"])
        dv = hv - lv; ds = "+" if dv > 0 else ""
        lines.append(f"| {r['agent']} | {lv:.1f}/5 | {hv:.0f}/5 | {ds}{dv:.1f} |")
    lines.append("")

    # ── Per-sample reviewer notes ─────────────────────────────────────────────
    lines += ["---\n", "## 4. Reviewer Notes\n"]
    for r in scored:
        notes = (r.get("human_notes") or "").strip()
        if notes:
            lines.append(f"**{r['agent']}:** {notes}\n")

    # ── Detail records backing Predict / Simulate (50 individually-reviewed
    #    records each) — capped at DETAIL_DISPLAY_CAP rows here; the agent-level
    #    row above is their mean, cross-checked against that mean on every run.
    lines += ["---\n", "## 5. Detailed Per-Record Reviews (Predict & Simulate)\n"]
    for agent_name, sheet_name in _DETAIL_SHEETS.items():
        try:
            detail_rows = _load_detail_sheet(sheet_name)
        except KeyError:
            continue
        if not detail_rows:
            continue
        n = len(detail_rows)
        avg = {
            d: sum(float(r[f"human_{d}"]) for r in detail_rows) / n
            for d in DIMENSIONS
        }

        summary_row = next((r for r in scored if r["agent"] == agent_name), None)
        check_line = ""
        if summary_row:
            summary_vals = {d: float(summary_row[f"human_{d}"]) for d in DIMENSIONS}
            matches = all(abs(summary_vals[d] - avg[d]) < 0.01 for d in DIMENSIONS)
            check_line = (
                f"Average across all {n} records — relevance {avg['relevance']:.2f}, "
                f"faithfulness {avg['faithfulness']:.2f}, safety {avg['safety']:.2f} — "
                f"{'matches' if matches else '**does NOT match**'} the {agent_name} row "
                f"in the Summary table above.\n"
            )

        lines += [
            f"### {agent_name} — {n} individually-reviewed records\n",
            check_line,
            f"Showing first {min(DETAIL_DISPLAY_CAP, n)} of {n}. Full set: "
            f"`evals/human_baseline/human_scores.xlsx` → `{sheet_name}` sheet.\n",
            "| Delivery ID | Human Relevance | Human Faithfulness | Human Safety | Mean |",
            "|---:|:---:|:---:|:---:|---:|",
        ]
        for r in detail_rows[:DETAIL_DISPLAY_CAP]:
            rel, fai, saf = float(r["human_relevance"]), float(r["human_faithfulness"]), float(r["human_safety"])
            lines.append(
                f"| {r.get('delivery_id', '')} | {rel:.0f}/5 | {fai:.0f}/5 | {saf:.0f}/5 "
                f"| **{(rel + fai + saf) / 3:.2f}** |"
            )
        lines.append("")

    report_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nHuman baseline report written: {report_path}")
