"""
Compare LLM-as-judge scores against human baseline scores.

Usage:
    uv run python evals/llm_judge_eval.py

Fill in human_relevance, human_faithfulness, human_safety (1-5) in the
"human_scores" sheet of evals/human_baseline/human_scores.xlsx, then run
this script. (Predict / Simulate additionally have 50-record detail sheets —
see test_eval_human_baseline.py, which is the full pytest-based version of
this comparison and covers those too. This script is the quick console-only
check against just the 5-row agent-level summary.)
"""

import sys
from pathlib import Path

import openpyxl

BASELINE_XLSX = Path(__file__).parent / "human_baseline" / "human_scores.xlsx"
DIMENSIONS    = ["relevance", "faithfulness", "safety"]


def load_scores(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["human_scores"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows, skipped = [], []
    for i, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
        vals = [c.value for c in row_cells]
        if all(v is None for v in vals):
            continue
        row = {headers[j]: str(vals[j]).strip() if vals[j] not in (None, "") else "" for j in range(len(headers))}
        if all((row.get(f"human_{d}") or "").strip() for d in DIMENSIONS):
            rows.append(row)
        else:
            skipped.append(row.get("sample_id", f"row{i}"))
    return rows, skipped


def main() -> None:
    if not BASELINE_XLSX.exists():
        print(f"XLSX not found: {BASELINE_XLSX}")
        sys.exit(1)

    rows, skipped = load_scores(BASELINE_XLSX)

    if skipped:
        print(f"Skipped (human scores missing): {', '.join(skipped)}")

    if not rows:
        print(
            "No rows have human scores filled in yet.\n"
            f"Open {BASELINE_XLSX} in Excel and fill in human_relevance, "
            "human_faithfulness, human_safety (1-5) for each sample."
        )
        sys.exit(0)

    print(f"\nHuman baseline comparison — {len(rows)} scored sample(s)\n")
    print(f"{'Agent':<35} {'LLM Mean':>9}  {'Human Mean':>10}  {'Diff':>6}  {'Aligned'}")
    print("-" * 75)

    for r in rows:
        llm_mean   = sum(float(r[f"llm_{d}"])   for d in DIMENSIONS) / 3
        human_mean = sum(float(r[f"human_{d}"]) for d in DIMENSIONS) / 3
        diff       = human_mean - llm_mean
        sign       = "+" if diff > 0 else ""
        aligned    = "Yes" if abs(diff) <= 1.0 else "No — review"
        print(f"  {r['agent']:<33} {llm_mean:>9.2f}  {human_mean:>10.2f}  {sign}{diff:>5.2f}  {aligned}")

    print()
    for d in DIMENSIONS:
        llm_vals   = [float(r[f"llm_{d}"])   for r in rows]
        human_vals = [float(r[f"human_{d}"]) for r in rows]
        avg_diff   = sum(h - l for l, h in zip(llm_vals, human_vals)) / len(rows)
        sign = "+" if avg_diff > 0 else ""
        print(f"  {d.capitalize():<15}  avg diff: {sign}{avg_diff:.2f}")


if __name__ == "__main__":
    main()
